import openpyxl
# I learned about how to use openpyxl from free online book "Automate the
# Boring Stuff With Python" (https://automatetheboringstuff.com/chapter12/)


# opens the workbooks and gets the sheets needed
wb2010 = openpyxl.load_workbook('DL_Dashboard_AY2010_2011_Q1.xlsx')
wb2015 = openpyxl.load_workbook('DL_Dashboard_AY2015_2016_Q1.xlsx')
sheet2010 = wb2010.get_sheet_by_name('DL_Dashboard_AY2010_2011_Q1')
sheet2015 = wb2015.get_sheet_by_name('DL_Dashboard_AY2015_2016_Q1')

clean2010 = []
clean2015 = []

# takes the necessary data and puts it into a list so it's easier to work with
# in python
def clean_data(sheet, clean_list_name):

    for i in range(7,4925): # this goes to end of 2010 data and past end of 2015
        try:
            # converts zip code from long type to int type
            # gets last digit of zip so can exclude all zips ending in 3, 5, 7
            zipcode = int(sheet.cell(row=i, column=4).value)
            last_dig = zipcode % 10
        except TypeError:
            # for 2015 goes past the end and raises TypeError because NoneType
            # can't be converted to int so skips these
            next
        except ValueError:
            # some foreign schools have zips with letters and raises ValueError
            # so skips these
            next

        try:
            # converts school type from unicode to python string and makes all
            # letters lowercase so it's easier to work with
            school_type = (sheet.cell(row=i, column=5).value).encode('utf-8').lower()
        except AttributeError:
            # for 2015 goes past the end and raises TypeError because empty
            # cells are NoneType and can't be converted so skips these
            next

        if last_dig != 3 and last_dig != 5 and last_dig != 7 and 'foreign' not in school_type:
            # selects only rows that meet the criteria (excludes zips ending in
            # 3, 5, 7 and all foreign schools)
            school_name = sheet.cell(row=i, column=2).value
            data = [school_name, zipcode, school_type]

            # selects all values from # of recipients to end for each row and
            # adds it to the 'data' list
            for j in range(6,36):
                try:
                    loan = int(sheet.cell(row=i, column=j).value)
                except ValueError:
                    # for cells with '-' raises ValueError so assigns these to 0
                    loan = 0
                except TypeError:
                    # 2015 sheet has fewer columns so goes past the end and
                    # empty cells raise TypeError so skips these
                    next

                data.append(loan)

            # adds 'data' list (with school name, zip, school type, and all loan
            # values to large list of all sheet info so now all data is stored
            # in large list where each element is another list with all info for
            # one school
            clean_list_name.append(data)


clean_data(sheet2010, clean2010)
clean_data(sheet2015, clean2015)

# Answering the questions

def question_the_first(dataset1, dataset2):
    """
    Consider all of the schools that disbursed a total of greater than or equal
    to $2000 and less than $9500 in loans for the time period reported on the
    spreadsheets provided (i.e. quarter 1). How many more schools met this
    criteria in 2015 than in 2010?
    """

    dataset1_count = 0
    dataset2_count = 0

    for school in dataset1:
        # iterates through each school and adds amount of disbursed loans for
        # each category (undergraduate, graduate, subsidized, unsubsidized, etc)
        total = school[7] + school[12] + school[17] + school[22] + school[27] + school[32]
        # if total is greater than or equal to 2000 and less than 9500 it adds
        # one to the count for dataset1 (when called will be 2010 data)
        if total >= 2000 and total < 9500:
            dataset1_count += 1

    for school in dataset2:
        # for 2015 data, there is one less category so fewer elements to add
        total = school[7] + school[12] + school[17] + school[22] + school[27]
        if total >= 2000 and total < 9500:
            dataset2_count += 1

    # finds how many more schools in 2015 than in 2010 meet the criteria
    difference = dataset2_count - dataset1_count

    print "Number of schools that disbursed >= $2000 and < $9500 in loans"
    print "in 2010-2011 Q1:"
    print dataset1_count
    print "Number of schools that disbursed >= $2000 and < $9500 in loans"
    print "in 2015-2016 Q1:"
    print dataset2_count
    print "How many more in 2015 than in 2010:"
    print difference


question_the_first(clean2010, clean2015)


def question_the_second(dataset1, dataset2):
    """
    Consider the sum of expected total loan amount if the loans were fully
    disbursed for each school. In 2010 as compared to 2015, for how many more
    schools was this amount greater than $20,000,000?
    """
    # The question is a bit ambiguous because it's not clear if they mean to
    # consider the expected total loan amount as if the loans were fully
    # disbursed, or only consider those loans which were fully disbursed. I ran
    # it both ways, and out of the schools in the lists only 1 in 2015 fully
    # disbursed the loans and none in 2010, so I left it as the first way
    # (considering the expected total loan amount as if they were fully disb)
    # because that seemed to make more sense.

    dataset1_count = 0
    dataset2_count = 0

    for school in dataset1:
        total = school[5] + school[10] + school[15] + school[20] + school[25] + school[30]
        if total > 20000000:
            dataset1_count += 1

    for school in dataset2:
        total = school[5] + school[10] + school[15] + school[20] + school[25]
        if total > 20000000:
            dataset2_count += 1

    difference = dataset1_count - dataset2_count

    print "Number of schools with > $20,000,000 originated loans in 2010 Q1:"
    print dataset1_count
    print "Number of schools with > $20,000,000 originated loans in 2015 Q1:"
    print dataset2_count
    print "How many more in 2010 than in 2015:"
    print difference


question_the_second(clean2010, clean2015)


def question_the_third(dataset):
    """
    In 2015, consider all the colonial colleges founded before the declaration
    of independence. Amongst these, what was the largest number of recipients
    within a school for either DL Graduate or DL Grad Plus loans?
    """

    # list of colonial colleges; source: wikipedia
    colonial_colleges = [
        u'HARVARD UNIVERSITY', u'COLLEGE OF WILLIAM AND MARY',
        u'YALE UNIVERSITY', u'PRINCETON UNIVERSITY',
        u'COLLUMBIA UNIVERSITY IN THE CITY OF NEW YORK', u'UNIVERSITY OF PENNSYLVANIA',
        u'BROWN UNIVERSITY', u'RUTGERS, THE STATE UNIVERSITY OF NEW JERSEY',
        u'DARTMOUTH COLLEGE']

    # creates a subset of data for schools that are in the colonial colleges list
    subset = []

    for school in dataset:
        if school[0] in colonial_colleges:
            subset.append(school)

    most_recipients = None

    for school in subset:
        # checks # of recipients for both DL graduate unsubsidized and DL grad
        # plus (grad subsidized is combined with undergrad subsidized so didn't
        # consider those) and selects the highest # of recipients from either
        if school[13] > school[23]:
            high = school[13]
        else:
            high = school[23]

        # compares highest # of recipients from current school to previous
        # highest (or assigns current as highest if it's the first) and
        # re-assigns most_recipients variable if current school is higher
        if most_recipients == None or high > most_recipients:
            most_recipients = high

    print "Most recipients for either DL Grad (unsubsidized)"
    print "or DL Grad Plus loans out of colonial colleges in 2015:"
    print most_recipients


question_the_third(clean2015)


def question_the_fourth(dataset):
    """
    In 2015, consider all the private nonprofit schools in King county, WA. For
    these schools, consider the expected total loan amount if the loan is fully
    disbursed for unsubsidized graduate studies. Exclude all schools where the
    unsubsidized graduate loan amount is not available i.e. "-" or 0. What was
    the median value?
    """
    # This q (like q 2) is ambiguous about the "if the loan is fully disbursed"
    # condition. I also ran this one both ways and if you only consider schools
    # with loans fully disbursed (out of private nonprofit schools in King
    # county WA), there are none, so I decided that can't be right and left it
    # the first way (considering expected total loan amount).

    # list of zip codes in King county WA; source: http://www.ciclt.net
    king_county_zips = [
        98001, 98002, 98003, 98004, 98005, 98006, 98007, 98008, 98008, 98010,
        98011, 98013, 98014, 98015, 98019, 98022, 98023, 98024, 98025, 98027,
        98028, 98029, 98030, 98031, 98032, 98033, 98034, 98035, 98038, 98039,
        98040, 98041, 98042, 98045, 98047, 98050, 98051, 98052, 98053, 98054,
        98055, 98056, 98057, 98058, 98059, 98062, 98063, 98064, 98065, 98068,
        98070, 98071, 98072, 98073, 98074, 98075, 98083, 98092, 98093, 98101,
        98102, 98103, 98104, 98105, 98106, 98107, 98108, 98109, 98111, 98112,
        98114, 98115, 98116, 98117, 98118, 98119, 98121, 98122, 98124, 98125,
        98126, 98131, 98132, 98133, 98134, 98136, 98138, 98144, 98145, 98146,
        98148, 98154, 98155, 98158, 98160, 98161, 98164, 98166, 98168, 98171,
        98174, 98177, 98178, 98188, 98198, 98199, 98224, 98288
    ]

    # creates a subset of data for schools that have a zip in King county WA
    subset = []

    for school in dataset:
        if school[1] in king_county_zips and school[2] == 'private-nonprofit':
            subset.append(school)

    # from the data subset, creates a list of expected amount of unsubsidized
    # graduate loans (only adds value to the list if value > 0 so it excludes
    # schools where amount was unavailable or 0)
    unsub_grad_loans = []

    for school in subset:
        if school[15] > 0:
            unsub_grad_loans.append(school[15])

    # puts the list in order from smallest to biggest and then finds the median
    # (if there's an even number of values, it averages the middle two)
    unsub_grad_loans.sort()

    if len(unsub_grad_loans) % 2 == 1:
        idx = len(unsub_grad_loans) / 2
        median = unsub_grad_loans[idx]
    else:
        idx = len(unsub_grad_loans) / 2
        median = (unsub_grad_loans[idx] + unsub_grad_loans[idx - 1]) / 2.0

    print "Private nonprofit schools in King County, WA:"
    for school in subset:
        print school[0]

    print "Median value of unsubsidized graduate loans for private nonprofit"
    print "schools in King County, WA in 2015 (if loans were > 0):"
    print median


question_the_fourth(clean2015)


def question_the_fifth(dataset):
    """
    Consider all the schools who were football champions in National Collegiate
    Athletic Association (reference: www.ncaa.com) from 2001 to 2015 (both years
    included) in the FBS division. For these schools (only include main campuses
    that participate in football and not all campuses) consider the expected
    total loan amount if the loan is fully disbursed for GRAD PLUS loans in 2015
    data provided. Calculate the sum of these amounts.
    """
    # This q also has the "if the loan is fully disbursed" condition, and like
    # the others I ran it both ways and there were no schools that met the
    # criteria that had loans fully disbursed, so I again left it the first way.

    ncaa_champions = [
        u'ALABAMA STATE UNIVERSITY', u'OHIO STATE UNIVERSITY',
        u'FLORIDA STATE UNIVERSITY', u'AUBURN UNIVERSITY',
        u'UNIVERSITY OF FLORIDA',
        u'LOUISIANA STATE UNIVERSITY & AGRICULTURAL & MECHANICAL COLLEGE',
        u'UNIVERSITY OF TEXAS - AUSTIN', u'UNIVERSITY OF SOUTHERN CALIFORNIA',
        u'UNIVERSITY OF MIAMI'
    ]

    # creates a subset of data for schools in the ncaa_champions list
    subset = []

    for school in dataset:
        if school[0] in ncaa_champions:
            subset.append(school)

    # adds values for expected total loan amount for Grad Plus loans for schools
    # in the data subset
    loan_sum = 0

    for school in subset:
        loan_sum += school[25]

    print "NCAA champion schools (2001-2015):"
    for school in subset:
        print school[0]
    print "Total amount of Grad Plus loans in 2015 for NCAA champion schools"
    print "(2001-2015):"
    print loan_sum


question_the_fifth(clean2015)
